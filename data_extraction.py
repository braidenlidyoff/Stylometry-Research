'''This program is meant to analyze everything from p.39 and on, of the Old English Corpus
It stores values like WorkID, CollectionID, Title, and Text in a DataFrame
The dataframe on output is truncated, and can be fixed to show the entire body of text'''

import os
from openpyxl import load_workbook
import pandas as pd
import pdfplumber
import re

def save_current_work(current_work, current_collection, current_title, current_text, output_folder, data):
    """Save the current work's text to a file and append its info to the data list."""
    if current_work and current_text:
        # Extract just the letter-number portion of current_work (e.g., "A1.1")
        file_work = current_work.split(":")[0].strip()
        current_text_filename = f"{file_work}.txt"
        text_filepath = os.path.join(output_folder, current_text_filename)
        try:
            with open(text_filepath, 'w', encoding='utf-8') as f:
                f.write("\n".join(current_text))
            print(f"Text saved to {text_filepath}")
        except Exception as e:
            print(f"Error saving text to {text_filepath}: {e}")
        data.append([current_work, current_collection, current_title, current_text_filename])

def data_extract(pdf_path, start_page, end_page, output_folder):
    data = []
    current_work = None
    current_collection = None
    current_title = None
    current_text = []
    
    # Ensure the output folder exists
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num in range(start_page - 1, end_page):
            page = pdf.pages[page_num]
            text = page.extract_text() or ""  # Ensures we always get a string
            lines = text.split("\n")  # Splits text into lines

            for line in lines:
                line = line.strip()  # Clean up whitespace

                # Skip unwanted lines
                if re.match(r'^\d{4}\. Bounds, Sawyer \d+ \(DOE Electronic edition\)', line): 
                    continue  # Skip lines like "1256. Bounds, Sawyer 635 (DOE Electronic edition)"

                if line in ["THE OLD ENGLISH CORPUS", "A: Texts"]:  # Skip section headers
                    continue

                # Detect WorkID (e.g., "B1.1.1: GenA,B")
                if re.match(r'^[A-Z]\d+(\.\d+)*:\s*\S+', line):
                    # Save previous work
                    save_current_work(current_work, current_collection, current_title, current_text, output_folder, data)
                    
                    current_work = line.strip()
                    # Extract CollectionID (e.g., "B1" from "B1.1.1")
                    current_collection = re.match(r'^([A-Z]\d+)', current_work).group(1)
                    current_text = []  # Reset text collection
                    continue

                # Extract Title
                if line.startswith("Title:"):
                    current_title = line.replace("Title:", "").strip().split(":")[0]
                    continue

                # Skip citations
                if line.startswith("Citation:"):
                    continue

                # Skip the numbers, but keep the content
                line = re.sub(r'^\d+:\s*', '', line)  # Remove line numbers
                current_text.append(line)
            
        # Save last collected work
        save_current_work(current_work, current_collection, current_title, current_text, output_folder, data)

    # Convert data to DataFrame
    df = pd.DataFrame(data, columns=["WorkID", "CollectionID", "Title", "Text File Name"])
    return df

# Path & Function Call
pdf_path = r"/Users/braidenlidyoff/Stylometry-Research/The Old English Corpus.pdf"
output_folder = "output_text_files"
df = data_extract(pdf_path, start_page=2740, end_page=2900, output_folder=output_folder)

output_file = "parsed_old_english_corpus.xlsx"

if os.path.exists(output_file):
    # Load the workbook and increment sheet number
    book = load_workbook(output_file)
    new_sheet_number = len(book.sheetnames) + 1
    sheet_name = f"Sheet{new_sheet_number}"

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
else:
    # If file does not exist, create new file with Sheet1
    sheet_name = "Sheet1"
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

print(f"Data saved to {output_file}, sheet: {sheet_name}")
